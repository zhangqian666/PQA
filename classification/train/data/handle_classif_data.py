# -*- coding: utf-8 -*-

"""
@author: zhangqian

@contact: 

@Created on: 2020-08-05 10:19
"""
# import pandas as pd
#
# print(pd.read_csv('./end.tsv', delimiter='t'))



from openpyxl import load_workbook
import csv

tag_list = ["UpAndDownSentences_simple", "CottonWadOrder_simple", "FillInTheWords_simple"
    , "ErrorCorrection_simple", "Disorder_simple", "Other_simple", "TheMeaningOfWords_multi"]

workbook = load_workbook("/Users/zhangqian/PycharmProjects/PQA/EL/data/train_classif/3000.xlsx")  # 找到需要xlsx文件的位置
booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet

# 获取sheet页的行数据
rows = booksheet.rows
# 获取sheet页的列数据
columns = booksheet.columns

end_data = []

i = 0
# 迭代所有的行
for row in rows:
    i = i + 1
    line = [col.value for col in row]
    cell_data_1 = booksheet.cell(row=i, column=2).value  # 获取第i行1 列的数据
    if cell_data_1 in tag_list:
        cell_data_2 = booksheet.cell(row=i, column=1).value  # 获取第i行 2 列的数据
        end_data.append([cell_data_1, cell_data_2])
        print([cell_data_1, cell_data_2])

with open("./end3.tsv", "w") as f:
    tsv_w = csv.writer(f)
    tsv_w.writerow(['type', 'string'])
    for i in end_data:
        tsv_w.writerow(i)
